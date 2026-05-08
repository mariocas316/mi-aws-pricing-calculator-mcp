#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter

# URLs regeneradas
urls = {
    'On-Demand': 'https://calculator.aws/#/estimate?id=b85917e49169b0323353e144b2365f7505e02edf',
    'SP 1yr': 'https://calculator.aws/#/estimate?id=3212057dc73dacbe78827859f33ad41ea4a53e91',
    'SP 3yr': 'https://calculator.aws/#/estimate?id=62d931b4d78ff55624c6524b6651be2cf599b0b7'
}

# Cargar Excel
wb = openpyxl.load_workbook(r'archivostcoonnet\TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')
ws = wb.active

# Fila 8 es AKS Nodes
row = 8

# Encontrar la última columna con datos
max_col = ws.max_column
print(f"Última columna usada: {max_col} ({get_column_letter(max_col)})")

# Ver las últimas 5 columnas
print("\nÚltimas 5 columnas de la fila 8:")
for col in range(max(1, max_col - 4), max_col + 1):
    cell = ws.cell(row, col)
    print(f"  Col {col} ({get_column_letter(col)}): {cell.value}")

# Agregar las URLs después de la última columna usada
# Voy a poner las URLs en las columnas siguientes
next_col = max_col + 1

# Agregar títulos en la fila anterior (fila 7) si es necesario
if ws.cell(row - 1, next_col).value is None:
    ws.cell(row - 1, next_col).value = "On-Demand URL"
    ws.cell(row - 1, next_col + 1).value = "SP 1yr URL"
    ws.cell(row - 1, next_col + 2).value = "SP 3yr URL"

# Agregar las URLs
ws.cell(row, next_col).value = urls['On-Demand']
ws.cell(row, next_col + 1).value = urls['SP 1yr']
ws.cell(row, next_col + 2).value = urls['SP 3yr']

print(f"\n✓ URLs agregadas en columnas {next_col} ({get_column_letter(next_col)}) a {next_col + 2} ({get_column_letter(next_col + 2)})")

# Guardar
wb.save(r'archivostcoonnet\TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')
print("✓ Archivo guardado")
