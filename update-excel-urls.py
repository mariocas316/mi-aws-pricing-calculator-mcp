#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter

# URLs finales con 10 EKS control planes (uno por cluster)
urls = {
    'On-Demand': 'https://calculator.aws/#/estimate?id=b5b5b809a4f18478695e2e77fe06452d6c02735d',
    'SP 1yr': 'https://calculator.aws/#/estimate?id=83d7d2208ae21a01e025052e9961cd4d82a11941',
    'SP 3yr': 'https://calculator.aws/#/estimate?id=9f4758759a7befa7163ad0bb6d368004cb457f73'
}

# Cargar Excel
wb = openpyxl.load_workbook(r'archivostcoonnet\TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')
ws = wb.active

row = 8  # AKS Nodes
header_row = 7

# Columnas para URLs (N, O, P = 14, 15, 16)
cols = {
    'On-Demand': 14,
    'SP 1yr': 15,
    'SP 3yr': 16
}

# Agregar encabezados en fila 7
ws.cell(header_row, cols['On-Demand']).value = 'Calculator On-Demand'
ws.cell(header_row, cols['SP 1yr']).value = 'Calculator SP 1yr'
ws.cell(header_row, cols['SP 3yr']).value = 'Calculator SP 3yr'

# Agregar URLs en fila 8
ws.cell(row, cols['On-Demand']).value = urls['On-Demand']
ws.cell(row, cols['SP 1yr']).value = urls['SP 1yr']
ws.cell(row, cols['SP 3yr']).value = urls['SP 3yr']

print("✓ Encabezados y URLs agregados:")
print(f"  Fila 7 (Encabezados):")
print(f"    Col N: {ws.cell(header_row, 14).value}")
print(f"    Col O: {ws.cell(header_row, 15).value}")
print(f"    Col P: {ws.cell(header_row, 16).value}")

print(f"\n  Fila 8 (URLs):")
print(f"    Col N: {ws.cell(row, 14).value}")
print(f"    Col O: {ws.cell(row, 15).value}")
print(f"    Col P: {ws.cell(row, 16).value}")

# Guardar
wb.save(r'archivostcoonnet\TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')
print("\n✓ Archivo actualizado y guardado")
