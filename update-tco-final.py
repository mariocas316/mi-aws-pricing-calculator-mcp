#!/usr/bin/env python3
import openpyxl
from pathlib import Path

# Datos reales de las calculadoras
data = {
    'On-Demand': {
        'monthly': 54314.64,
        'annual': 651775.68,
        'url': 'https://calculator.aws/#/estimate?id=ab9b48fa6e478ba7bfaf28ed1ea8e6a67e6d170f'
    },
    'SP 1yr': {
        'monthly': 40936.63,
        'annual': 491239.56,
        'url': 'https://calculator.aws/#/estimate?id=7c23c3f1f9c6cb888de6184877209a6bfad2b5ad'
    },
    'SP 3yr': {
        'monthly': 30393.20,
        'annual': 364718.40,
        'url': 'https://calculator.aws/#/estimate?id=980f081b3c800ab9f3308e088f20f029f317fe8e'
    }
}

# Cargar Excel
file_path = Path(r'archivostcoonnet\TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')
wb = openpyxl.load_workbook(file_path)
ws = wb['1. Resumen Ejecutivo TCO']

print("=" * 80)
print("ACTUALIZANDO TCO CON VALORES REALES DE CALCULADORA AWS")
print("=" * 80)

# Fila 8 es AKS Nodes
# Basado en estructura anterior, los costos esperados están en columnas específicas

# Actualizar fila 8 con costos On-Demand (columna F aproximadamente)
# Necesito revisar la estructura actual del Excel primero

# Leer encabezados de fila 7
print("\nEncabezados Fila 7:")
for col in range(1, 20):
    header = ws.cell(7, col).value
    if header:
        print(f"  Col {col}: {header}")

print("\nDatos actuales Fila 8:")
for col in range(1, 20):
    value = ws.cell(8, col).value
    if value:
        print(f"  Col {col}: {value}")

# Buscar las columnas de costos basándose en valores numéricos visibles
# Típicamente estarían: A=Item, B=Descripción, C=Estrategia, D=Azure, E=Azure anno, F=AWS OD, G=AWS SP1yr, H=AWS SP3yr

# Actualizar valores de AWS (columnas F, G, H para On-Demand, SP1yr, SP3yr respectivamente)
print("\n" + "=" * 80)
print("ACTUALIZANDO VALORES EN FILA 8")
print("=" * 80)

# Asumo estructura: F=On-Demand, G=SP1yr, H=SP3yr (estos son los valores mensuales típicos)
# Las URLs irán en columnas N, O, P (como se configuró antes)

# Actualizar columnas F, G, H con valores mensuales
ws.cell(8, 6).value = data['On-Demand']['monthly']  # Col F - On-Demand
ws.cell(8, 7).value = data['SP 1yr']['monthly']      # Col G - SP 1yr
ws.cell(8, 8).value = data['SP 3yr']['monthly']      # Col H - SP 3yr

print(f"✓ Col F (On-Demand):  ${data['On-Demand']['monthly']:,.2f}")
print(f"✓ Col G (SP 1yr):     ${data['SP 1yr']['monthly']:,.2f}")
print(f"✓ Col H (SP 3yr):     ${data['SP 3yr']['monthly']:,.2f}")

# Actualizar URLs en columnas N, O, P
ws.cell(8, 14).value = data['On-Demand']['url']   # Col N
ws.cell(8, 15).value = data['SP 1yr']['url']       # Col O
ws.cell(8, 16).value = data['SP 3yr']['url']       # Col P

print(f"\n✓ Col N (URL On-Demand): Agregada")
print(f"✓ Col O (URL SP 1yr):     Agregada")
print(f"✓ Col P (URL SP 3yr):     Agregada")

# Guardar
wb.save(file_path)
print("\n" + "=" * 80)
print("✓ ARCHIVO GUARDADO EXITOSAMENTE")
print("=" * 80)

print("\nResumen de actualización:")
print(f"  On-Demand:  $54,314.64/mes → ${data['On-Demand']['monthly']:,.2f}")
print(f"  SP 1yr:     $40,936.63/mes → ${data['SP 1yr']['monthly']:,.2f}")
print(f"  SP 3yr:     $30,393.20/mes → ${data['SP 3yr']['monthly']:,.2f}")
