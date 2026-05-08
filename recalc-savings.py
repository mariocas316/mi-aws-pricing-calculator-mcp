#!/usr/bin/env python3
import openpyxl
from pathlib import Path

# Datos reales
azure_monthly = 54472
aws_ondemand = 54314.64
aws_sp1yr = 40936.63
aws_sp3yr = 30393.20

# Calculadora ahorro basado en SP 3yr (el escenario más optimista/recomendado)
ahorro_mensual_sp3yr = azure_monthly - aws_sp3yr
ahorro_anual_sp3yr = ahorro_mensual_sp3yr * 12
ahorro_pct_sp3yr = (ahorro_mensual_sp3yr / azure_monthly) * 100 if azure_monthly > 0 else 0

print("=" * 80)
print("RECALCULANDO AHORROS CON VALORES REALES")
print("=" * 80)

print(f"\nComparación (Azure vs AWS SP 3yr):")
print(f"  Azure (actual):        ${azure_monthly:>10,.2f}/mes")
print(f"  AWS SP 3yr (propuesto): ${aws_sp3yr:>10,.2f}/mes")
print(f"  Ahorro mensual:         ${ahorro_mensual_sp3yr:>10,.2f}/mes")
print(f"  Ahorro %:               {ahorro_pct_sp3yr:>10.1f}%")
print(f"  Ahorro anual:           ${ahorro_anual_sp3yr:>10,.2f}/año")

# Actualizar Excel
file_path = Path(r'archivostcoonnet\TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')
wb = openpyxl.load_workbook(file_path)
ws = wb['1. Resumen Ejecutivo TCO']

# Fila 8, actualizar:
# Col 9: Ahorro Mensual (SP 3yr) = Azure - SP3yr
# Col 10: Ahorro % = Ahorro/Azure
# Col 11: Ahorro Anual = Ahorro Mensual * 12

ws.cell(8, 9).value = ahorro_mensual_sp3yr        # Col I - Ahorro Mensual
ws.cell(8, 10).value = ahorro_pct_sp3yr / 100     # Col J - Ahorro % (como decimal)
ws.cell(8, 11).value = ahorro_anual_sp3yr         # Col K - Ahorro Anual

print("\n" + "=" * 80)
print("VALORES ACTUALIZADOS EN FILA 8")
print("=" * 80)
print(f"✓ Col I (Ahorro Mensual):  ${ahorro_mensual_sp3yr:,.2f}")
print(f"✓ Col J (Ahorro %):        {ahorro_pct_sp3yr:.1f}%")
print(f"✓ Col K (Ahorro Anual):    ${ahorro_anual_sp3yr:,.2f}")

# Guardar
wb.save(file_path)

print("\n" + "=" * 80)
print("✓ TCO ACTUALIZADO COMPLETAMENTE")
print("=" * 80)

print(f"\nRESUMEN FINAL - Fila 8 (AKS Nodes):")
print(f"  Azure Mensual:           ${azure_monthly:>10,.2f}")
print(f"  AWS On-Demand:           ${aws_ondemand:>10,.2f}")
print(f"  AWS SP 1yr:              ${aws_sp1yr:>10,.2f}")
print(f"  AWS SP 3yr (recomendado):${aws_sp3yr:>10,.2f}")
print(f"  ---")
print(f"  Ahorro mensual (SP 3yr): ${ahorro_mensual_sp3yr:>10,.2f} ({ahorro_pct_sp3yr:.1f}%)")
print(f"  Ahorro anual:            ${ahorro_anual_sp3yr:>10,.2f}")
print(f"\nURLs de calculadora agregadas en columnas N, O, P")
