#!/usr/bin/env python3
"""
Actualiza la fila de APIM en Excel con costos de la calculadora
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

excel_file = Path('C:/Users/mario.castillo/proyectos/mi-aws-pricing-calculator-mcp/archivostcoonnet/TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')

# Costos de la calculadora proporcionados por el usuario
COSTO_ONDEMAND = 404.84
COSTO_SP1YR = 404.84
COSTO_SP3YR = 404.84

print("=" * 80)
print("ACTUALIZANDO COSTOS DE APIM EN EXCEL")
print("=" * 80)

try:
    # Leer el archivo
    df = pd.read_excel(excel_file, sheet_name='1. Resumen Ejecutivo TCO')
    
    print(f"\n📖 Archivo leído: {excel_file}")
    print(f"   Total de filas: {len(df)}")
    
    # Buscar la fila de APIM
    print("\n🔍 Buscando fila de APIM...")
    apim_row_index = None
    
    for idx, row in df.iterrows():
        # Buscar en la primera columna
        cell_value = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
        if 'apim' in cell_value:
            apim_row_index = idx
            print(f"✓ Encontrada en fila {idx + 2} (Excel row {idx + 1})")
            print(f"  Contenido: {row.iloc[0]}")
            break
    
    if apim_row_index is None:
        print("❌ No se encontró fila de APIM")
        print("\nPrimeras 15 filas para referencia:")
        print(df.iloc[:15, :3])
    else:
        print(f"\n✏️  Actualizando costos en fila APIM (índice {apim_row_index}):")
        
        # Actualizar costos en columnas F, G, H (índices 5, 6, 7)
        df.iloc[apim_row_index, 5] = COSTO_ONDEMAND   # Col F: On-Demand
        df.iloc[apim_row_index, 6] = COSTO_SP1YR      # Col G: SP 1yr
        df.iloc[apim_row_index, 7] = COSTO_SP3YR      # Col H: SP 3yr
        
        print(f"  • Col F (On-Demand): ${COSTO_ONDEMAND:.2f}")
        print(f"  • Col G (SP 1yr): ${COSTO_SP1YR:.2f}")
        print(f"  • Col H (SP 3yr): ${COSTO_SP3YR:.2f}")
        
        # Recalcular ahorros si existen valores Azure
        azure_monthly = df.iloc[apim_row_index, 3]  # Col D
        if pd.notna(azure_monthly) and isinstance(azure_monthly, (int, float)):
            ahorro_mensual = azure_monthly - COSTO_SP3YR
            ahorro_porcentaje = (ahorro_mensual / azure_monthly) if azure_monthly > 0 else 0
            ahorro_anual = ahorro_mensual * 12
            
            df.iloc[apim_row_index, 8] = round(ahorro_mensual, 2)      # Col I
            df.iloc[apim_row_index, 9] = round(ahorro_porcentaje, 6)   # Col J
            df.iloc[apim_row_index, 10] = round(ahorro_anual, 2)       # Col K
            
            print(f"\n  Ahorros recalculados (vs Azure ${azure_monthly:.2f}):")
            print(f"  • Col I (Ahorro Mensual): ${ahorro_mensual:.2f}")
            print(f"  • Col J (Ahorro %): {ahorro_porcentaje*100:.1f}%")
            print(f"  • Col K (Ahorro Anual): ${ahorro_anual:.2f}")
        
        # Guardar con openpyxl para preservar formato
        wb = load_workbook(excel_file)
        ws = wb['1. Resumen Ejecutivo TCO']
        
        # Actualizar celdas
        excel_row = apim_row_index + 2  # +1 para convertir a 1-indexed, +1 más para header
        ws[f'F{excel_row}'] = COSTO_ONDEMAND
        ws[f'G{excel_row}'] = COSTO_SP1YR
        ws[f'H{excel_row}'] = COSTO_SP3YR
        
        if pd.notna(azure_monthly) and isinstance(azure_monthly, (int, float)):
            ws[f'I{excel_row}'] = round(ahorro_mensual, 2)
            ws[f'J{excel_row}'] = round(ahorro_porcentaje, 6)
            ws[f'K{excel_row}'] = round(ahorro_anual, 2)
        
        wb.save(excel_file)
        print(f"\n✅ Excel actualizado y guardado")

except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
