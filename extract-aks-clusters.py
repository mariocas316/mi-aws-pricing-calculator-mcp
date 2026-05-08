#!/usr/bin/env python3
"""
Extrae información de clusters AKS del inventario de Azure
y mapea a homólogos en AWS
"""
import openpyxl
from pathlib import Path
from collections import defaultdict

inventory_file = Path('inventarioAri/AzureResourceInventory_Report_2026-04-16_16_43.xlsx')

if not inventory_file.exists():
    print(f"✗ Archivo no encontrado")
    exit(1)

wb = openpyxl.load_workbook(inventory_file)

print("=" * 100)
print("ANÁLISIS DE CLUSTERS AKS - INVENTARIO AZURE")
print("=" * 100)

# Investigar todas las hojas
print("\n📋 Hojas del inventario:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Buscar información de "AKS" o "Container"
print("\n" + "=" * 100)
print("BUSCANDO INFORMACIÓN DE AKS/CLUSTERS")
print("=" * 100)

aks_data = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Buscar todas las filas que contengan información relevante
    for row_idx in range(1, sheet.max_row + 1):
        row_data = []
        has_aks_keyword = False
        
        for col_idx in range(1, min(sheet.max_column + 1, 20)):  # Primeras 20 columnas
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            
            if value:
                row_data.append(str(value))
                if 'AKS' in str(value).upper() or 'KUBERNETES' in str(value).upper() or 'CLUSTER' in str(value).upper():
                    has_aks_keyword = True
        
        if has_aks_keyword and row_data:
            key = f"{sheet_name}#{row_idx}"
            aks_data[key] = row_data
            print(f"\n🔍 Hoja: {sheet_name} | Fila: {row_idx}")
            for idx, val in enumerate(row_data, 1):
                print(f"   Col {idx}: {val}")

# Buscar información de instancias/máquinas
print("\n" + "=" * 100)
print("BUSCANDO INSTANCIAS/MÁQUINAS EN CLUSTERS")
print("=" * 100)

vm_patterns = ['virtual machine', 'vm', 'instance', 'node', 'sku', 'size']

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    print(f"\n📄 {sheet_name}")
    
    # Si tiene columnas que sugieran datos de VMs
    if sheet.max_row > 1:
        # Leer encabezados
        headers = []
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            header_cell = sheet.cell(1, col_idx)
            if header_cell.value:
                headers.append((col_idx, str(header_cell.value)))
        
        if headers:
            print(f"   Encabezados encontrados:")
            for col_idx, header in headers:
                print(f"      Col {col_idx}: {header}")
        
        # Mostrar primeras 10 filas de datos
        print(f"   Datos (primeras 5 filas):")
        data_rows = 0
        for row_idx in range(2, min(sheet.max_row + 1, 7)):
            row_values = []
            for col_idx, header in headers:
                cell = sheet.cell(row_idx, col_idx)
                if cell.value:
                    row_values.append(f"{cell.value}")
            
            if row_values:
                print(f"      Fila {row_idx}: {' | '.join(row_values[:5])}")  # Primeras 5 columnas
                data_rows += 1

print("\n" + "=" * 100)
print("RESUMEN")
print("=" * 100)
print(f"Total de hojas: {len(wb.sheetnames)}")
print(f"Registros AKS encontrados: {len(aks_data)}")
