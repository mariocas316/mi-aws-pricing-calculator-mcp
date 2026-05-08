#!/usr/bin/env python3
import openpyxl
from pathlib import Path

# Leer el inventario de Azure
inventory_file = Path('inventarioAri/AzureResourceInventory_Report_2026-04-16_16_43.xlsx')

if not inventory_file.exists():
    print(f"✗ Archivo no encontrado: {inventory_file}")
    exit(1)

wb = openpyxl.load_workbook(inventory_file)

print("=" * 80)
print("ANÁLISIS DEL INVENTARIO AZURE - CLUSTERS AKS")
print("=" * 80)
print(f"\n📋 Hojas disponibles en el libro:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Buscar información de AKS
print("\n" + "=" * 80)
print("EXTRAYENDO INFORMACIÓN DE AKS")
print("=" * 80)

aks_clusters = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Buscar filas con "AKS" en cualquier columna
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = str(row).upper()
        if 'AKS' in row_str or 'KUBERNETES' in row_str:
            print(f"\n📍 Hoja: {sheet_name} | Fila: {row_idx}")
            print(f"   Contenido: {row}")

# Buscar "ContainerInstances" o similares
print("\n" + "=" * 80)
print("BUSCANDO INSTANCIAS DE CONTENEDOR")
print("=" * 80)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Leer todas las filas
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), 1):
        if row[0] is None:
            continue
        
        row_str = str(row[0]).upper() if row[0] else ""
        
        if any(keyword in row_str for keyword in ['CONTAINER', 'NODE', 'VM', 'MACHINE', 'INSTANCE', 'CLUSTER', 'KUBERNETES']):
            print(f"\n🔍 Hoja: {sheet_name} | Fila: {row_idx}")
            for col_idx, cell_value in enumerate(row, 1):
                if cell_value is not None:
                    print(f"   Col {col_idx}: {cell_value}")

print("\n" + "=" * 80)
print("RESUMEN DE HOJAS")
print("=" * 80)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n📄 {sheet_name}")
    print(f"   Filas: {sheet.max_row} | Columnas: {sheet.max_column}")
    
    # Mostrar primeras filas de encabezado
    print(f"   Encabezados:")
    for col_idx, cell in enumerate(sheet[1], 1):
        if cell.value:
            print(f"      Col {col_idx}: {cell.value}")
