#!/usr/bin/env python3
"""
Calcula costos de API Gateway para 80M llamadas mensuales
Y actualiza Excel con los costos - VERSIÓN CORREGIDA CON RUTA CORRECTA
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Ruta correcta del archivo Excel
excel_file = Path('C:/Users/mario.castillo/proyectos/onnet/TCO_Migracion_Azure_AWS_OnNet_v2.xlsx')

# Precios AWS API Gateway
API_CALLS_MONTHLY = 80_000_000  # 80 millones
API_COST_PER_MILLION = 3.50

# Data Transfer aproximada
DATA_TRANSFER_GB = 100
DATA_TRANSFER_COST_PER_GB = 0.09

# CloudWatch Logs
LOGS_GB = 50
LOGS_COST_PER_GB = 0.50

# Calcular costos
print("=" * 80)
print("ANÁLISIS DE COSTOS - API GATEWAY (80M llamadas/mes)")
print("=" * 80)

# Desglose
api_calls_millions = API_CALLS_MONTHLY / 1_000_000
api_cost_monthly = (api_calls_millions * API_COST_PER_MILLION)
data_transfer_cost_monthly = DATA_TRANSFER_GB * DATA_TRANSFER_COST_PER_GB
logs_cost_monthly = LOGS_GB * LOGS_COST_PER_GB
total_monthly_ondemand = api_cost_monthly + data_transfer_cost_monthly + logs_cost_monthly

print(f"\n💰 COSTOS MENSUALES (On-Demand):")
print(f"   • API Calls ({api_calls_millions}M × ${API_COST_PER_MILLION}/M):  ${api_cost_monthly:,.2f}")
print(f"   • Data Transfer ({DATA_TRANSFER_GB}GB × ${DATA_TRANSFER_COST_PER_GB}/GB):    ${data_transfer_cost_monthly:,.2f}")
print(f"   • CloudWatch Logs ({LOGS_GB}GB × ${LOGS_COST_PER_GB}/GB):   ${logs_cost_monthly:,.2f}")
print(f"   {'─' * 50}")
print(f"   • TOTAL On-Demand:                 ${total_monthly_ondemand:,.2f}")

# Savings Plans
sp1yr_discount = 0.25
sp3yr_discount = 0.45

cost_sp1yr_monthly = total_monthly_ondemand * (1 - sp1yr_discount)
cost_sp3yr_monthly = total_monthly_ondemand * (1 - sp3yr_discount)

print(f"\n📊 ESCENARIOS CON SAVINGS PLANS:")
print(f"   • SP 1yr (25% desc):              ${cost_sp1yr_monthly:,.2f}/mes")
print(f"   • SP 3yr (45% desc):              ${cost_sp3yr_monthly:,.2f}/mes")

# Anuales
annual_ondemand = total_monthly_ondemand * 12
annual_sp1yr = cost_sp1yr_monthly * 12
annual_sp3yr = cost_sp3yr_monthly * 12

print(f"\n📈 COSTOS ANUALES:")
print(f"   • On-Demand:                      ${annual_ondemand:,.2f}")
print(f"   • SP 1yr:                         ${annual_sp1yr:,.2f}")
print(f"   • SP 3yr:                         ${annual_sp3yr:,.2f}")

# Actualizar Excel usando openpyxl
print(f"\n📝 Actualizando Excel: {excel_file}")

try:
    wb = load_workbook(excel_file)
    ws = wb['1. Resumen Ejecutivo TCO']
    
    # Encontrar última fila con datos
    last_row = ws.max_row
    print(f"   Última fila con datos: {last_row}")
    
    # APIM será la siguiente fila después de AKS (fila 8 + 1 = fila 9)
    apim_row = 9  # Excel row (1-indexed)
    
    print(f"\n✏️  Actualizando fila {apim_row}:")
    
    # Información básica (Columnas A-C)
    ws[f'A{apim_row}'] = 'APIM (11 instancias, 80M llamadas/mes)'
    ws[f'B{apim_row}'] = 'Amazon API Gateway'
    ws[f'C{apim_row}'] = 'Replatform'
    
    print(f"  • Col A: APIM (11 instancias, 80M llamadas/mes)")
    print(f"  • Col B: Amazon API Gateway")
    print(f"  • Col C: Replatform")
    
    # Costos Azure (aproximado)
    azure_monthly = 1100
    azure_annual = azure_monthly * 12
    
    ws[f'D{apim_row}'] = azure_monthly      # Azure Mensual
    ws[f'E{apim_row}'] = azure_annual       # Azure Anual
    
    print(f"  • Col D (Azure Monthly): ${azure_monthly:,.2f}")
    print(f"  • Col E (Azure Annual): ${azure_annual:,.2f}")
    
    # Costos AWS (Columnas F-H)
    ws[f'F{apim_row}'] = round(total_monthly_ondemand, 2)     # On-Demand
    ws[f'G{apim_row}'] = round(cost_sp1yr_monthly, 2)         # SP 1yr
    ws[f'H{apim_row}'] = round(cost_sp3yr_monthly, 2)         # SP 3yr
    
    print(f"  • Col F (AWS On-Demand): ${total_monthly_ondemand:,.2f}")
    print(f"  • Col G (AWS SP 1yr): ${cost_sp1yr_monthly:,.2f}")
    print(f"  • Col H (AWS SP 3yr): ${cost_sp3yr_monthly:,.2f}")
    
    # Ahorros (Columnas I-K)
    ahorro_mensual = total_monthly_ondemand - cost_sp3yr_monthly
    ahorro_porcentaje = (ahorro_mensual / total_monthly_ondemand) * 100 if total_monthly_ondemand > 0 else 0
    ahorro_anual = ahorro_mensual * 12
    
    ws[f'I{apim_row}'] = round(ahorro_mensual, 2)              # Ahorro Mensual
    ws[f'J{apim_row}'] = round(ahorro_porcentaje / 100, 6)    # Ahorro %
    ws[f'K{apim_row}'] = round(ahorro_anual, 2)               # Ahorro Anual
    
    print(f"  • Col I (Ahorro Mensual): ${ahorro_mensual:,.2f}")
    print(f"  • Col J (Ahorro %): {ahorro_porcentaje:.1f}%")
    print(f"  • Col K (Ahorro Anual): ${ahorro_anual:,.2f}")
    
    # URLs de calculadora (Columnas N-P)
    ws[f'N{apim_row}'] = 'https://calculator.aws/#/estimate?id=17a2b57f568036aee836395e441cd6849d45177a'
    ws[f'O{apim_row}'] = 'https://calculator.aws/#/estimate?id=1963f6eb5befc2b1ec267ab873d93a41a22dcb47'
    ws[f'P{apim_row}'] = 'https://calculator.aws/#/estimate?id=8373dd6d9be6d6ec0a78ef690239be8ae99d43d7'
    
    print(f"\n📎 URLs de calculadora agregadas:")
    print(f"  • Col N (On-Demand)")
    print(f"  • Col O (SP 1yr)")
    print(f"  • Col P (SP 3yr)")
    
    # Guardar
    wb.save(excel_file)
    print(f"\n✅ Excel actualizado exitosamente")
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
print("RESUMEN FINAL - APIM → API GATEWAY")
print("=" * 80)
print(f"Instancias: 11 APIM")
print(f"Llamadas: 80 millones mensuales")
print(f"Costos AWS:")
print(f"  • On-Demand:  ${total_monthly_ondemand:>8,.2f}/mes  (${annual_ondemand:>12,.2f}/año)")
print(f"  • SP 1yr:     ${cost_sp1yr_monthly:>8,.2f}/mes  (${annual_sp1yr:>12,.2f}/año)")
print(f"  • SP 3yr:     ${cost_sp3yr_monthly:>8,.2f}/mes  (${annual_sp3yr:>12,.2f}/año) ← RECOMENDADO")
print("=" * 80)
